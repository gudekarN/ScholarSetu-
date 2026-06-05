from app.utils import get_academic_year
from flask_mail import Message
from app.models import Students
from datetime import datetime
from flask import jsonify
from app.extensions import db, mail
from app.models import InviteTokens, Notices, Queries, QueryReplies, StudentScholarshipData
from flask import abort
from app.models import College
from flask import render_template
from flask import request
from flask import Blueprint
from flask import session
from flask import redirect
from flask import url_for
from flask import make_response
import uuid 
import openpyxl
from io import BytesIO
from flask import send_file

admin_bp=Blueprint('admin', __name__, url_prefix='/admin')

# Admin dashboard

@admin_bp.route("/dashboard")
def dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        session.clear() # wipe any partial/wrong session
        return redirect(url_for('auth.login'))

    college = College.query.filter_by(college_id=session['user_id']).first()
    
    # Handle invalid or missing college safely
    if not college:
        session.clear()
        return redirect(url_for('auth.login'))

    initials = "NA"
    if college.college_name:
        words = [w for w in college.college_name.replace('.', ' ').split() if w]
        if words:
            if len(words) == 1:
                initials = words[0][:2].upper()
            elif words[0].isupper() and len(words[0]) >= 2:
                initials = words[0][:2].upper()
            else:
                initials = (words[0][0] + words[1][0]).upper()

    response = make_response(render_template("admin_dashboard.html", college=college, initials=initials))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma']        = 'no-cache'
    response.headers['Expires']       = '0'
    return response

# Generate token (Generate by admin)

@admin_bp.route("/generate_token", methods=["GET", "POST"])
def generate_token():
    if request.method=="POST":
        if 'user_id' not in session or session["role"]!="admin":
            return abort(401)

        existing=InviteTokens.query.filter_by(college_id=session['user_id'], is_active=True).first()

        if existing:
            existing.is_active=False
            db.session.commit()

        token=str(uuid.uuid4())

        new_token=InviteTokens(
            college_id=session['user_id'],
            token=token,
            is_active=True
        )

        try:
            db.session.add(new_token)
            db.session.commit()
            return jsonify({"success": True, "token": token})

        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False})

# Get Token (This code is define for get the real token show in generate token field for the perticular college)

@admin_bp.route("/get_token")
def get_token():
    if 'user_id' not in session or session.get('role') != 'admin':
        return abort(401)

    inviteToken=InviteTokens.query.filter_by(college_id=session['user_id'], is_active=True).first()

    if inviteToken:
        token=inviteToken.token
        createdAt=inviteToken.created_at

        student_count = Students.query.filter_by(college_id=session['user_id']).count()

        return jsonify({"success":True, "token":token, "created_at":str(createdAt), "student_count":student_count})

    return jsonify({"success":False})

@admin_bp.route("/post_notice", methods=["POST"])
def post_notice():
    if request.method=="POST":

        if 'user_id' not in session or session["role"]!="admin":
            return abort(401)

        data=request.get_json()

        title=data.get("title")
        content=data.get("content")
        notice_type =data.get("type")
        deadline_date=data.get("deadline_date")

        parsed_deadline = None
        if deadline_date:
            parsed_deadline = datetime.strptime(deadline_date, "%Y-%m-%d")

        db_college=College.query.filter_by(college_id=session["user_id"]).first()

        if db_college:

            notice=Notices(
                college_id=session["user_id"],
                title= title,
                content= content,
                type= notice_type ,
                deadline_date= parsed_deadline,
                academic_year=get_academic_year()
            )

            try:
                db.session.add(notice)
                db.session.commit()

                if notice_type =="high":
                    db_student=Students.query.filter_by(college_id=session["user_id"]).all()

                    emails = [s.email for s in db_student]

                    html_body = f"""
                    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: auto; padding: 20px; border: 1px solid #eee;">
                        <h2 style="color: #e07b00;">{title}</h2>
                        <p style="font-size: 15px; color: #333;">{content}</p>
                        <hr style="border: none; border-top: 1px solid #eee;">
                        <p style="font-size: 12px; color: #888;">This is an automated notice from ScholarSetu. Do not reply to this email.</p>
                    </div>
                    """

                    message = Message(
                        subject=title,
                        sender="gudekarnihar96@gmail.com",
                        recipients=emails
                    )
                    message.html = html_body
                    mail.send(message)

                return jsonify({"success":True})
                
            except Exception as e:
                db.session.rollback()
                return jsonify({"success":False})
        
        return jsonify({"success":False})

@admin_bp.route("/get_notices")
def get_notices():
    if 'user_id' not in session or session["role"]!="admin":
        return abort(401)

    notices=Notices.query.filter_by(college_id=session["user_id"]).order_by(Notices.posted_at.desc()).all()

    if notices:

        data=[]

        for notice in notices:
            data.append({
                "notice_id": notice.notice_id,
                "title": notice.title,
                "content": notice.content,
                "type": notice.type,
                "deadline_date": str(notice.deadline_date) if notice.deadline_date else None,
                "posted_at": notice.posted_at.strftime("%d %b %Y") if notice.posted_at else None
            })

        return jsonify({"success": True, "notices": data})

    return jsonify({"success":False})

@admin_bp.route("/get_students")
def get_students():
    if "user_id" not in session or session["role"]!="admin":
        return abort(401)

    db_students=Students.query.filter_by(college_id=session["user_id"]).order_by(Students.student_id.asc()).all()

    if db_students:
        data=[]

        import json as _json
        for student in db_students:
            scholarship = StudentScholarshipData.query.filter_by(
                student_id=student.student_id,
                college_id=session['user_id']
            ).first()

            extra = {}
            if scholarship and scholarship.extra_data:
                try:
                    extra = _json.loads(scholarship.extra_data)
                except Exception:
                    extra = {}

            data.append({
                "student_id": student.student_id,
                "full_name": student.full_name,
                "prn": student.prn,
                "email": student.email,
                "contact_number": student.contact_number,
                "department": student.department,
                "year": student.year,
                "is_verified": student.is_verified,
                "joined_at": datetime.strftime(student.joined_at, "%Y-%m-%d"),
                "extra_data": extra
            })

        return jsonify({"success":True, "students":data})

    return jsonify({"success":False})
        
@admin_bp.route("/update_student", methods=["POST"])
def update_student():
    if request.method=="POST":
        if "user_id" not in session or session["role"]!="admin":
            return abort(401)

        data=request.get_json()
        student_id= data.get("student_id")

        db_student=Students.query.filter_by(college_id=session["user_id"], student_id=student_id).first()

        if db_student:
            try:
                db_student.full_name= data.get("full_name", db_student.full_name)
                db_student.prn= data.get("prn", db_student.prn)
                db_student.email= data.get("email", db_student.email)
                db_student.contact_number= data.get("contact_number", db_student.contact_number)
                db_student.department= data.get("department", db_student.department)
                db_student.year= data.get("year", db_student.year)

                extra_data = data.get("extra_data")  # dict or None
                if extra_data is not None:
                    record = StudentScholarshipData.query.filter_by(
                        student_id=student_id,
                        college_id=session['user_id']
                    ).first()
                    if record:
                        import json as _json
                        old = _json.loads(record.extra_data) if record.extra_data else {}
                        old.update(extra_data)
                        record.extra_data = _json.dumps(old)

                db.session.commit()

                return jsonify({"success":True})

            except Exception as e:
                db.session.rollback()
                
        return jsonify({"success":False}) 

@admin_bp.route("/get_queries")
def get_queries():
    if "user_id" not in session or session["role"]!="admin":
            return abort(401)

    db_queries=Queries.query.filter_by(college_id=session["user_id"]).order_by(Queries.posted_at.desc()).all()

    queries=[]
    
    for query in db_queries:
        db_student=Students.query.filter_by(student_id=query.student_id).first()

        db_queryReplies=QueryReplies.query.filter_by(query_id=query.query_id).order_by(QueryReplies.replied_at.asc()).all()

        queryReplies=[]

        for reply in db_queryReplies:
            if reply.is_admin:
                responder_name = "Admin"
                responder_email = "Replay By Admin" if db_student else ""
            else:
                responder = Students.query.filter_by(student_id=reply.responder_id).first()
                responder_name = responder.full_name if responder else "Student"
                responder_email = responder.email if responder else ""

            queryReplies.append({
                "reply_id":        reply.reply_id,
                "is_admin":        reply.is_admin,
                "reply_text":      reply.reply_text,
                "replied_at":      str(reply.replied_at),
                "responder_name":  responder_name,
                "responder_email": responder_email
            })

        queries.append({
            "query_id": query.query_id,
            "question_text": query.question_text,
            "replies": queryReplies,
            "is_answered": query.is_answered,
            "posted_at": str(query.posted_at),
            "student_name": db_student.full_name,
            "student_email": db_student.email
        })

    return jsonify({"success": True, "questions": queries})

@admin_bp.route("/post_reply", methods=["POST"])
def post_reply():
    if "user_id" not in session or session["role"]!="admin":
        return abort(401)

    data=request.get_json()

    query_id=data.get("query_id")
    reply_text=data.get("reply_text")

    if not reply_text or len(reply_text.strip()) <=2:
        return jsonify({"success": False})

    query=Queries.query.filter_by(college_id=session["user_id"], query_id=query_id).first()    
    
    if not query:
        return jsonify({"success":False})

    queryreply=QueryReplies(
        query_id=query_id, 
        responder_id=session['user_id'], 
        is_admin=True, 
        reply_text=reply_text
    )

    query=Queries.query.get(query_id)
    query.is_answered=True

    try:
        db.session.add(queryreply)
        db.session.commit()
        
        return jsonify({"success":True})
    except Exception as e:
        db.session.rollback()

        return jsonify({"success":False})

@admin_bp.route("/generate_excel", methods=["GET"])
def generate_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        return abort(401)

    try:
        custom_cols = request.args.getlist('cols')

        students = Students.query.filter_by(
            college_id=session['user_id']
        ).order_by(Students.student_id.asc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        from openpyxl.styles import PatternFill, Font, Protection
        from openpyxl.utils import get_column_letter

        fixed_headers = ["PRN", "FULL NAME"]
        all_headers = fixed_headers + (custom_cols if custom_cols else [
            "Application ID", "Scheme Name", "Category", "Status"
        ])

        locked_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        locked_font = Font(bold=True, color="000000")

        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if header in fixed_headers:
                cell.fill = locked_fill
                cell.font = locked_font
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

        for row_idx, student in enumerate(students, start=2):
            prn_cell = ws.cell(row=row_idx, column=1, value=student.prn)
            name_cell = ws.cell(row=row_idx, column=2, value=student.full_name)
            prn_cell.protection = Protection(locked=True)
            name_cell.protection = Protection(locked=True)

            # Explicitly unlock all custom column cells for this row
            for col_idx in range(3, len(all_headers) + 1):
                ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=False)

        ws.protection.sheet = True
        ws.protection.password = "scholarsetu"
        ws.protection.enable()

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        for i in range(3, len(all_headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"students_{session['user_id']}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "trace": traceback.format_exc()}), 500

@admin_bp.route("/parse_excel", methods=["POST"])
def parse_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        return abort(401)

    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({"success": False, "message": "Please upload a valid .xlsx file"}), 400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception:
        return jsonify({"success": False, "message": "Could not read the Excel file"}), 400

    success_count = 0
    unmatched = []
    skipped = []
    rows = []  # full row data for preview table

    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

    if 'PRN' not in headers:
        return jsonify({"success": False, "message": "Excel file is missing required PRN column"}), 400

    prn_col = headers.index('PRN')

    skip_cols = {
        'prn', 'full name', 'email', 'contact number', 'contact',
        'department', 'year', 'verified', 'joined at'
    }
    custom_col_indices = [
        (i, headers[i]) for i in range(len(headers))
        if headers[i].strip().lower() not in skip_cols and headers[i].strip() != ''
    ]

    for row in ws.iter_rows(min_row=2, values_only=True):
        prn = str(row[prn_col]).strip() if row[prn_col] else None
        if not prn or prn == 'None':
            continue

        # Build a dict of all column values for the preview
        row_data = {headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))}

        student = Students.query.filter_by(
            prn=prn,
            college_id=session['user_id']
        ).first()

        if not student:
            unmatched.append(prn)
            row_data['_status'] = 'unmatched'
            rows.append(row_data)
            continue

        all_filled = all(
            row[i] is not None and str(row[i]).strip() not in ('', 'None')
            for i, _ in custom_col_indices
        )
        if not all_filled:
            skipped.append(prn)
            row_data['_status'] = 'skipped'
            rows.append(row_data)
            continue

        success_count += 1
        row_data['_status'] = 'success'
        rows.append(row_data)

    return jsonify({
        "success": True,
        "headers": [h for h in headers if h],
        "report": {
            "success_count": success_count,
            "total_rows": success_count + len(unmatched) + len(skipped),
            "unmatched": unmatched,
            "skipped": skipped,
            "rows": rows
        }
    })

@admin_bp.route("/confirm_excel", methods=["POST"])
def confirm_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        return abort(401)

    data = request.get_json()
    rows = data.get('rows', [])
    if not rows:
        return jsonify({"success": False, "message": "No rows provided to save"}), 400

    skip_cols = {
        'prn', 'full name', 'email', 'contact number', 'contact',
        'department', 'year', 'verified', 'joined at', '_status'
    }
    saved_count = 0

    import json as _json

    for row_data in rows:
        prn = str(row_data.get('PRN', '')).strip()
        if not prn or prn == 'None':
            continue

        student = Students.query.filter_by(
            prn=prn,
            college_id=session['user_id']
        ).first()

        if not student:
            continue

        all_filled = True
        extra = {}
        for col_name, val in row_data.items():
            if col_name.lower() not in skip_cols:
                if val is not None and str(val).strip() not in ('', 'None'):
                    extra[col_name] = str(val).strip()
                else:
                    all_filled = False

        if not all_filled or not extra:
            continue

        existing = StudentScholarshipData.query.filter_by(
            student_id=student.student_id,
            college_id=session['user_id']
        ).first()

        if existing:
            old_extra = _json.loads(existing.extra_data) if existing.extra_data else {}
            old_extra.update(extra)
            existing.extra_data = _json.dumps(old_extra)
            existing.added_via = "excel_upload"
        else:
            record = StudentScholarshipData(
                student_id=student.student_id,
                college_id=session['user_id'],
                application_id="",
                category="",
                scheme_name="",
                status="Pending",
                added_via="excel_upload",
                extra_data=_json.dumps(extra)
            )
            db.session.add(record)

        saved_count += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": "Database error during save"}), 500



    return jsonify({"success": True, "saved": saved_count})

@admin_bp.route("/export_data", methods=["GET"])
def export_data():
    if 'user_id' not in session or session.get('role') != 'admin':
        return abort(401)

    dept_filter = request.args.get('department')
    year_filter = request.args.get('year')

    query = Students.query.filter_by(college_id=session['user_id'])
    if dept_filter:
        query = query.filter_by(department=dept_filter)
    if year_filter:
        query = query.filter_by(year=year_filter)

    students = query.order_by(Students.student_id.asc()).all()

    import json as _json
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Data"

    # Collect all unique extra_data keys across all students
    extra_keys = []
    for student in students:
        scholarship = StudentScholarshipData.query.filter_by(
            student_id=student.student_id,
            college_id=session['user_id']
        ).first()
        if scholarship and scholarship.extra_data:
            try:
                extra = _json.loads(scholarship.extra_data)
                for k in extra.keys():
                    if k not in extra_keys:
                        extra_keys.append(k)
            except Exception:
                pass

    # Build fixed headers + dynamic extra headers
    fixed_headers = [
        "PRN", "FULL NAME", "EMAIL", "CONTACT",
        "DEPARTMENT", "YEAR", "VERIFIED", "JOINED AT"
    ]
    all_headers = fixed_headers + extra_keys

    # Style header row — saffron background
    header_fill = PatternFill(
        start_color="F97316", end_color="F97316", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write student rows — all cells editable (no protection)
    for row_idx, student in enumerate(students, start=2):
        scholarship = StudentScholarshipData.query.filter_by(
            student_id=student.student_id,
            college_id=session['user_id']
        ).first()

        extra = {}
        if scholarship and scholarship.extra_data:
            try:
                extra = _json.loads(scholarship.extra_data)
            except Exception:
                extra = {}

        ws.cell(row=row_idx, column=1, value=student.prn)
        ws.cell(row=row_idx, column=2, value=student.full_name)
        ws.cell(row=row_idx, column=3, value=student.email)
        ws.cell(row=row_idx, column=4, value=student.contact_number)
        ws.cell(row=row_idx, column=5, value=student.department)
        ws.cell(row=row_idx, column=6, value=student.year)
        ws.cell(row=row_idx, column=7, value="Yes" if student.is_verified else "No")
        ws.cell(row=row_idx, column=8, value=student.joined_at.strftime("%Y-%m-%d") if student.joined_at else "")

        for extra_idx, key in enumerate(extra_keys, start=len(fixed_headers) + 1):
            ws.cell(row=row_idx, column=extra_idx, value=extra.get(key, ""))

    # Set column widths
    col_widths = [20, 28, 32, 16, 22, 12, 10, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for i in range(len(fixed_headers) + 1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if dept_filter:
        label = f"Dept_{dept_filter}"
    elif year_filter:
        label = f"Year_{year_filter}"
    else:
        label = "Full_Report"

    college = College.query.filter_by(college_id=session['user_id']).first()
    college_name = college.college_name.replace(' ', '_') if college else str(session['user_id'])
    filename = f"{college_name}_{label}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@admin_bp.route("/delete_student", methods=["POST"])
def delete_student():
    if 'user_id' not in session or session.get('role') != 'admin':
        return abort(401)
    data = request.get_json()
    student_id = data.get('student_id')
    student = Students.query.filter_by(student_id=student_id, college_id=session['user_id']).first()
    if not student:
        return jsonify({"success": False})
    try:
        StudentScholarshipData.query.filter_by(student_id=student_id).delete()
        db.session.delete(student)
        db.session.commit()
        return jsonify({"success": True})
    except:
        db.session.rollback()
        return jsonify({"success": False})

@admin_bp.route("/delete_column", methods=["POST"])
def delete_column():
    if 'user_id' not in session or session.get('role') != 'admin':
        return abort(401)
    data = request.get_json()
    col_name = data.get('column', '').strip()
    if not col_name:
        return jsonify({"success": False})
    import json as _json
    records = StudentScholarshipData.query.filter_by(college_id=session['user_id']).all()
    try:
        for record in records:
            if record.extra_data:
                extra = _json.loads(record.extra_data)
                if col_name in extra:
                    del extra[col_name]
                    record.extra_data = _json.dumps(extra)
        db.session.commit()
        return jsonify({"success": True})
    except:
        db.session.rollback()
        return jsonify({"success": False})